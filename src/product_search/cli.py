"""CLI 入口，基于 click + rich 提供美观的命令行界面。"""

import asyncio
import json
import sys
from pathlib import Path
from typing import Optional

import click
from rich.console import Console
from rich.panel import Panel
from rich.progress import (
    BarColumn,
    MofNCompleteColumn,
    Progress,
    SpinnerColumn,
    TaskProgressColumn,
    TextColumn,
    TimeElapsedColumn,
    TimeRemainingColumn,
)
from rich.table import Table

console = Console()


# ── 主命令组 ─────────────────────────────────────────────────────────────────

@click.group(invoke_without_command=True)
@click.version_option(version="0.1.0", prog_name="product-search")
@click.option(
    "--verbose", "-v",
    is_flag=True,
    default=False,
    help="显示详细工作流日志（默认静默，仅写入日志文件）",
)
@click.option(
    "--max-iterations", "-n",
    default=3,
    show_default=True,
    help="最大搜索迭代次数（REPL 模式下每次搜索生效）",
)
@click.option(
    "--llm-config",
    default="default",
    show_default=True,
    help="LLM 配置名称（对应 config.toml 中的 [llm.<name>]）",
)
@click.option(
    "--no-summary",
    is_flag=True,
    help="跳过 AI 汇总报告",
)
@click.pass_context
def main(ctx, verbose, max_iterations, llm_config, no_summary):
    """ProductSearch - AI 驱动的企业产品信息搜索工具。

    \b
    直接运行进入交互模式：
      product-search

    \b
    单次搜索：
      product-search search "示例科技"

    \b
    批量搜索：
      product-search batch 企业.xlsx
    """
    ctx.ensure_object(dict)
    ctx.obj["verbose"] = verbose
    ctx.obj["max_iterations"] = max_iterations
    ctx.obj["llm_config"] = llm_config
    ctx.obj["no_summary"] = no_summary

    from product_search.core.logger import set_stderr_level
    set_stderr_level("INFO" if verbose else "WARNING")

    # 无子命令时进入 REPL
    if ctx.invoked_subcommand is None:
        asyncio.run(_run_repl(max_iterations, llm_config, no_summary))


async def _run_repl(max_iterations: int, llm_config: str, no_summary: bool) -> None:
    from product_search.repl import run_repl
    await run_repl(max_iterations, llm_config, no_summary)


# ── search 命令 ───────────────────────────────────────────────────────────────

@main.command()
@click.argument("company_name")
@click.option(
    "--output", "-o",
    type=click.Choice(["table", "json", "text"]),
    default="table",
    help="输出格式：table（表格）、json（JSON）、text（纯文本）",
)
@click.option(
    "--max-iterations", "-n",
    default=3,
    show_default=True,
    help="最大搜索迭代次数",
)
@click.option(
    "--llm-config",
    default="default",
    show_default=True,
    help="LLM 配置名称（对应 config.toml 中的 [llm.<name>]）",
)
@click.option(
    "--no-summary",
    is_flag=True,
    help="跳过 AI 汇总报告，只输出产品列表",
)
@click.pass_context
def search(ctx, company_name: str, output: str, max_iterations: int, llm_config: str, no_summary: bool):
    """搜索指定企业的产品信息。

    COMPANY_NAME: 企业名称（如"示例科技"、"Acme Corp"）
    """
    asyncio.run(_run_search(company_name, output, max_iterations, llm_config, no_summary))


async def _run_search(
    company_name: str,
    output_format: str,
    max_iterations: int,
    llm_config: str,
    no_summary: bool,
):
    """异步执行搜索并输出结果。"""
    from product_search.graph.workflow import run_search
    from product_search.llm.factory import create_llm

    console.print(Panel.fit(
        f"[bold cyan]ProductSearch[/bold cyan]\n"
        f"企业：[bold]{company_name}[/bold]\n"
        f"最大迭代：{max_iterations} 次",
        title="搜索任务",
    ))

    try:
        llm = create_llm(llm_config)
        analysis_llm = None
        if not no_summary:
            try:
                analysis_llm = create_llm("analysis")
            except Exception:
                analysis_llm = llm
    except Exception as e:
        console.print(f"[red]LLM 初始化失败: {e}[/red]")
        console.print(
            "[yellow]提示：请检查 config/config.toml 中的 LLM 配置，"
            "或设置 OPENAI_API_KEY / ANTHROPIC_API_KEY 环境变量[/yellow]"
        )
        sys.exit(1)

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        TimeElapsedColumn(),
        console=console,
        transient=True,
    ) as progress:
        progress.add_task(f"正在搜索 [cyan]{company_name}[/cyan] 的产品信息...", total=None)
        try:
            state = await run_search(
                company_name=company_name,
                max_iterations=max_iterations,
                llm=llm,
                analysis_llm=analysis_llm if not no_summary else None,
            )
        except Exception as e:
            progress.stop()
            console.print(f"[red]搜索失败: {e}[/red]")
            sys.exit(1)

    products = state.get("products", [])
    summary = state.get("summary", "")

    if not products:
        console.print(f"[yellow]未找到关于 {company_name!r} 的产品信息。[/yellow]")
        console.print("建议：检查企业名称拼写，或尝试换用其他搜索引擎（修改 config.toml 中的 search.engine）")
        return

    if output_format == "json":
        _output_json(products, summary, company_name)
    elif output_format == "text":
        _output_text(products, summary, company_name)
    else:
        _output_table(products, summary, company_name)


# ── 输出格式函数 ──────────────────────────────────────────────────────────────

def _output_table(products, summary, company_name):
    """表格格式输出。"""
    console.print()
    table = Table(title=f"{company_name} 产品列表（共 {len(products)} 个）", show_lines=True)
    table.add_column("产品名称", style="bold cyan", min_width=20)
    table.add_column("类别", style="green", min_width=12)
    table.add_column("描述", min_width=30)
    table.add_column("置信度", justify="right", min_width=8)

    for p in products:
        conf_color = "green" if p.confidence >= 0.8 else "yellow" if p.confidence >= 0.6 else "red"
        table.add_row(
            p.name,
            p.category,
            p.description[:50] + "..." if len(p.description) > 50 else p.description,
            f"[{conf_color}]{p.confidence:.0%}[/{conf_color}]",
        )

    console.print(table)
    if summary:
        console.print()
        console.print(Panel(summary, title="AI 分析报告", border_style="blue"))


def _output_json(products, summary, company_name):
    """JSON 格式输出。"""
    output = {
        "company_name": company_name,
        "product_count": len(products),
        "products": [p.model_dump() for p in products],
        "summary": summary,
    }
    click.echo(json.dumps(output, ensure_ascii=False, indent=2))


def _output_text(products, summary, company_name):
    """纯文本格式输出。"""
    click.echo(f"\n{company_name} 产品列表（共 {len(products)} 个）")
    click.echo("=" * 50)
    for p in products:
        click.echo(f"• {p.name}（{p.category}）- 置信度: {p.confidence:.0%}")
        if p.description:
            click.echo(f"  {p.description}")
    if summary:
        click.echo("\nAI 分析报告：")
        click.echo("-" * 50)
        click.echo(summary)


# ── batch 命令 ────────────────────────────────────────────────────────────────

@main.command()
@click.argument("input_file", type=click.Path(exists=True, readable=True))
@click.option(
    "--output", "-o",
    default=None,
    help="输出 Excel 文件路径（默认：在输入文件同目录生成 products_output.xlsx）",
)
@click.option(
    "--company-column", "-c",
    default=None,
    help="企业名称所在列名或列号（从 1 开始），默认使用第一列",
)
@click.option(
    "--sheet",
    default=None,
    help="读取的工作表名称，默认使用第一个工作表",
)
@click.option(
    "--skip-header/--no-skip-header",
    default=True,
    show_default=True,
    help="是否跳过第一行表头",
)
@click.option(
    "--max-iterations", "-n",
    default=3,
    show_default=True,
    help="每家企业的最大搜索迭代次数",
)
@click.option(
    "--llm-config",
    default="default",
    show_default=True,
    help="LLM 配置名称（对应 config.toml 中的 [llm.<name>]）",
)
@click.option(
    "--no-summary",
    is_flag=True,
    help="跳过 AI 汇总报告",
)
@click.pass_context
def batch(
    ctx,
    input_file: str,
    output: Optional[str],
    company_column: Optional[str],
    sheet: Optional[str],
    skip_header: bool,
    max_iterations: int,
    llm_config: str,
    no_summary: bool,
):
    """从 Excel 批量读取企业名称，搜索每家企业的产品，输出到新 Excel 文件。

    INPUT_FILE: 输入的 Excel 文件路径（.xlsx / .xls）

    \b
    输入 Excel 格式：
      每行一家企业，默认读取第一列，支持有/无表头

    \b
    输出 Excel 列：
      企业名称 | 产品名称 | 产品类别 | 产品描述 | 置信度 | 来源URL | AI汇总
    """
    asyncio.run(_run_batch(
        input_file, output, company_column, sheet,
        skip_header, max_iterations, llm_config, no_summary,
    ))


async def _run_batch(
    input_file: str,
    output_path: Optional[str],
    company_column: Optional[str],
    sheet: Optional[str],
    skip_header: bool,
    max_iterations: int,
    llm_config: str,
    no_summary: bool,
):
    """异步执行批量搜索并写出 Excel。"""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        console.print("[red]缺少依赖 openpyxl，请执行：pip install openpyxl[/red]")
        sys.exit(1)

    from product_search.graph.workflow import run_search
    from product_search.llm.factory import create_llm

    # ── 读取输入 Excel ──────────────────────────────────────────────
    input_path = Path(input_file).resolve()
    try:
        wb_in = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    except Exception as e:
        console.print(f"[red]无法读取 Excel 文件: {e}[/red]")
        sys.exit(1)

    ws_in = wb_in[sheet] if sheet else wb_in.active

    col_idx = 1
    if company_column:
        if company_column.isdigit():
            col_idx = int(company_column)
        else:
            header_row = next(ws_in.iter_rows(min_row=1, max_row=1, values_only=True), ())
            normalized = [str(h).strip() if h else "" for h in header_row]
            if company_column in normalized:
                col_idx = normalized.index(company_column) + 1
            else:
                console.print(f"[red]未找到列名 '{company_column}'，可用列：{normalized}[/red]")
                sys.exit(1)

    start_row = 2 if skip_header else 1
    companies = []
    for row in ws_in.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx, values_only=True):
        val = row[0]
        if val is not None and str(val).strip():
            companies.append(str(val).strip())
    wb_in.close()

    if not companies:
        console.print("[yellow]输入文件中未找到任何企业名称，请检查列设置和表头选项。[/yellow]")
        sys.exit(1)

    console.print(
        f"\n读取到 [bold cyan]{len(companies)}[/bold cyan] 家企业，"
        f"每家最多搜索 [bold]{max_iterations}[/bold] 轮，开始搜索...\n"
    )

    # ── 初始化 LLM ──────────────────────────────────────────────────
    try:
        llm = create_llm(llm_config)
        analysis_llm = None
        if not no_summary:
            try:
                analysis_llm = create_llm("analysis")
            except Exception:
                analysis_llm = llm
    except Exception as e:
        console.print(f"[red]LLM 初始化失败: {e}[/red]")
        sys.exit(1)

    # ── 批量搜索（带进度条）────────────────────────────────────────
    all_results = []

    with Progress(
        SpinnerColumn(),
        TextColumn("[bold blue]{task.description}"),
        BarColumn(),
        MofNCompleteColumn(),
        TaskProgressColumn(),
        TimeElapsedColumn(),
        TimeRemainingColumn(),
        console=console,
        transient=False,
    ) as progress:
        task = progress.add_task("批量搜索企业产品", total=len(companies))

        for company in companies:
            progress.update(task, description=f"搜索中：[cyan]{company}[/cyan]")
            try:
                state = await run_search(
                    company_name=company,
                    max_iterations=max_iterations,
                    llm=llm,
                    analysis_llm=analysis_llm if not no_summary else None,
                )
                products = state.get("products", [])
                summary = state.get("summary", "")
                all_results.append((company, products, summary))
                progress.console.print(
                    f"  [green]✓[/green] [bold]{company}[/bold]  "
                    f"[dim]找到 {len(products)} 个产品[/dim]"
                )
            except Exception as e:
                all_results.append((company, [], ""))
                progress.console.print(
                    f"  [red]✗[/red] [bold]{company}[/bold]  [red dim]{e}[/red dim]"
                )
            finally:
                progress.advance(task)

        progress.update(task, description="[green]搜索完成[/green]")

    # ── 生成输出 Excel ───────────────────────────────────────────────
    if output_path is None:
        output_path = str(input_path.parent / "products_output.xlsx")

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "产品列表"

    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["企业名称", "产品名称", "产品类别", "产品描述", "置信度", "来源URL", "AI汇总"]
    col_widths = [18, 22, 16, 45, 8, 40, 60]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    fill_colors = ["EBF3FB", "FFFFFF"]
    row_num = 2
    for ci, (company, products, summary) in enumerate(all_results):
        fill = PatternFill(fill_type="solid", fgColor=fill_colors[ci % 2])
        cell_align = Alignment(vertical="top", wrap_text=True)

        if not products:
            ws.cell(row=row_num, column=1, value=company)
            ws.cell(row=row_num, column=2, value="（未找到产品）")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = fill
                ws.cell(row=row_num, column=col).alignment = cell_align
            row_num += 1
            continue

        for pi, p in enumerate(products):
            ws.cell(row=row_num, column=1, value=company if pi == 0 else "")
            ws.cell(row=row_num, column=2, value=p.name)
            ws.cell(row=row_num, column=3, value=p.category)
            ws.cell(row=row_num, column=4, value=p.description)
            ws.cell(row=row_num, column=5, value=round(p.confidence, 2))
            ws.cell(row=row_num, column=6, value=p.source_url)
            ws.cell(row=row_num, column=7, value=summary if pi == 0 else "")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = fill
                ws.cell(row=row_num, column=col).alignment = cell_align
            row_num += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    ws_summary = wb_out.create_sheet("企业汇总")
    sum_headers = ["企业名称", "产品数量", "产品类别数", "AI汇总"]
    sum_widths = [20, 10, 12, 80]
    for col, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws_summary.column_dimensions[get_column_letter(col)].width = w

    for ri, (company, products, summary) in enumerate(all_results, 2):
        categories = len({p.category for p in products})
        ws_summary.cell(row=ri, column=1, value=company)
        ws_summary.cell(row=ri, column=2, value=len(products))
        ws_summary.cell(row=ri, column=3, value=categories)
        ws_summary.cell(row=ri, column=4, value=summary)
        fill = PatternFill(fill_type="solid", fgColor=fill_colors[ri % 2])
        for col in range(1, len(sum_headers) + 1):
            ws_summary.cell(row=ri, column=col).fill = fill
            ws_summary.cell(row=ri, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    try:
        wb_out.save(output_path)
    except Exception as e:
        console.print(f"[red]保存 Excel 失败: {e}[/red]")
        sys.exit(1)

    total_products = sum(len(p) for _, p, _ in all_results)
    console.print(
        f"\n[bold green]完成！[/bold green] "
        f"共搜索 [bold]{len(companies)}[/bold] 家企业，"
        f"找到 [bold]{total_products}[/bold] 个产品。"
    )
    console.print(f"输出文件：[bold cyan]{output_path}[/bold cyan]")


# ── llm 命令组 ────────────────────────────────────────────────────────────────

# 各 provider 对应的默认模型
_PROVIDER_DEFAULT_MODELS: dict[str, str] = {
    "openai":    "gpt-4o-mini",
    "anthropic": "claude-sonnet-4-5-20250929",
    "ollama":    "llama3.2",
    "deepseek":  "deepseek-chat",
    "glm":       "glm-4-flash",
    "minimax":   "MiniMax-Text-01",
    "kimi":      "moonshot-v1-8k",
    "qwen":      "qwen-turbo",
    "seed":      "",   # 需要显式传入接入点 ID
}

_KNOWN_PROVIDERS = set(_PROVIDER_DEFAULT_MODELS.keys())


@main.group()
def llm():
    """管理 LLM 模型配置（添加、查看、删除）。"""
    pass


@llm.command(name="add")
@click.argument("name")
@click.option("--provider", "-p", default=None,
              help="模型提供商（留空时从 NAME 推断）")
@click.option("--model", "-m", default=None,
              help="模型名称（留空时使用 provider 默认值）")
@click.option("--api-key", "-k", default="",
              help="API Key（也可通过环境变量设置，不填则留空）")
@click.option("--base-url", "-u", default="",
              help="自定义 API Base URL（留空使用内置默认地址）")
@click.option("--max-tokens", default=None, type=int,
              help="最大 token 数（默认 4096）")
@click.option("--temperature", default=None, type=float,
              help="采样温度（默认 0.7）")
def llm_add(name, provider, model, api_key, base_url, max_tokens, temperature):
    """添加或更新一个 LLM 配置块。

    NAME 为配置块名称，对应 config.toml 中的 [llm.NAME]。

    \b
    示例：
      product-search llm add deepseek --api-key sk-xxx
      product-search llm add kimi --model moonshot-v1-32k --api-key sk-xxx
      product-search llm add gpt4 --provider openai --model gpt-4o --api-key sk-xxx
      product-search llm add local --provider ollama --model qwen2.5
    """
    try:
        import tomlkit
    except ImportError:
        console.print("[red]缺少依赖 tomlkit，请执行：pip install tomlkit[/red]")
        sys.exit(1)

    from product_search.core.config import config
    from product_search.core.exceptions import ConfigError

    # 推断 provider（NAME 与 provider 同名时自动推断）
    if provider is None:
        if name in _KNOWN_PROVIDERS:
            provider = name
        else:
            console.print(
                f"[red]无法从名称 '{name}' 推断 provider，请通过 --provider 明确指定。[/red]\n"
                f"支持的 provider：{', '.join(sorted(_KNOWN_PROVIDERS))}"
            )
            sys.exit(1)

    provider = provider.lower()
    if provider not in _KNOWN_PROVIDERS:
        console.print(
            f"[yellow]警告：未知 provider '{provider}'，将按 OpenAI 兼容协议处理。[/yellow]"
        )

    # 确定模型名
    if model is None:
        model = _PROVIDER_DEFAULT_MODELS.get(provider, "")
        if not model:
            console.print(
                f"[red]Provider '{provider}' 无默认模型，"
                f"请通过 --model 指定（如豆包接入点 ID）。[/red]"
            )
            sys.exit(1)

    # 获取可写配置路径
    try:
        cfg_path = config.writable_config_path()
    except ConfigError as e:
        console.print(f"[red]{e}[/red]")
        sys.exit(1)

    # 读取现有配置（tomlkit 保留注释和格式）
    with open(cfg_path, "r", encoding="utf-8") as f:
        doc = tomlkit.load(f)

    if "llm" not in doc:
        doc.add("llm", tomlkit.table())

    llm_tbl = doc["llm"]
    is_update = name in llm_tbl

    # 构建字段（只写入显式传入的可选项）
    fields: dict = {"provider": provider, "model": model}
    if api_key:
        fields["api_key"] = api_key
    if base_url:
        fields["base_url"] = base_url
    if max_tokens is not None:
        fields["max_tokens"] = max_tokens
    if temperature is not None:
        fields["temperature"] = temperature

    if is_update:
        for k, v in fields.items():
            llm_tbl[name][k] = v
        action = "更新"
    else:
        section = tomlkit.table()
        for k, v in fields.items():
            section.add(k, v)
        llm_tbl.add(name, section)
        action = "添加"

    with open(cfg_path, "w", encoding="utf-8") as f:
        tomlkit.dump(doc, f)

    # 重新加载使当前进程感知到变更
    config._initialized = False
    config.__init__()

    console.print(f"\n[green]✓[/green] 已{action} LLM 配置：[bold cyan][llm.{name}][/bold cyan]")
    console.print(f"  provider    = [green]{provider}[/green]")
    console.print(f"  model       = [green]{model}[/green]")
    if api_key:
        masked = api_key[:6] + "***" if len(api_key) > 6 else "***"
        console.print(f"  api_key     = [dim]{masked}[/dim]")
    if base_url:
        console.print(f"  base_url    = [dim]{base_url}[/dim]")
    if max_tokens is not None:
        console.print(f"  max_tokens  = {max_tokens}")
    if temperature is not None:
        console.print(f"  temperature = {temperature}")
    console.print(f"\n配置文件：[dim]{cfg_path}[/dim]")
    console.print(
        f"使用示例：[dim]product-search search \"示例科技\" --llm-config {name}[/dim]"
    )


@llm.command(name="list")
def llm_list():
    """列出所有已配置的 LLM 模型。"""
    from product_search.core.config import _PROVIDER_ENV_KEYS, config

    llm_configs = config.llm
    if not llm_configs:
        console.print("[yellow]暂无 LLM 配置。[/yellow]")
        return

    table = Table(title="已配置的 LLM 模型", show_lines=True, border_style="blue")
    table.add_column("名称", style="bold cyan", min_width=12)
    table.add_column("Provider", style="green", min_width=10)
    table.add_column("模型", min_width=26)
    table.add_column("API Key", min_width=18)
    table.add_column("Base URL", min_width=20)

    for cfg_name, settings in llm_configs.items():
        # API Key 展示
        if settings.api_key:
            ak = settings.api_key
            ak_display = ak[:6] + "***" if len(ak) > 6 else "***"
        else:
            env_var = _PROVIDER_ENV_KEYS.get(settings.provider.lower(), "")
            if env_var and __import__("os").environ.get(env_var):
                ak_display = f"[dim]{env_var}[/dim]"
            else:
                ak_display = "[red]未配置[/red]"

        bu = settings.base_url if settings.base_url else "[dim]-[/dim]"
        table.add_row(cfg_name, settings.provider, settings.model, ak_display, bu)

    console.print()
    console.print(table)
    console.print()


@llm.command(name="remove")
@click.argument("name")
@click.option("--yes", "-y", is_flag=True, help="跳过确认提示")
def llm_remove(name, yes):
    """删除一个 LLM 配置块。

    NAME: 要删除的配置块名称（不能删除 default）
    """
    try:
        import tomlkit
    except ImportError:
        console.print("[red]缺少依赖 tomlkit，请执行：pip install tomlkit[/red]")
        sys.exit(1)

    from product_search.core.config import config
    from product_search.core.exceptions import ConfigError

    if name == "default":
        console.print("[red]不允许删除 default 配置块。[/red]")
        sys.exit(1)

    if not yes:
        click.confirm(f"确认删除 LLM 配置 '[bold]{name}[/bold]'？", abort=True)

    try:
        cfg_path = config.writable_config_path()
    except ConfigError as e:
        console.print(f"[red]{e}[/red]")
        sys.exit(1)

    with open(cfg_path, "r", encoding="utf-8") as f:
        doc = tomlkit.load(f)

    llm_tbl = doc.get("llm", {})
    if name not in llm_tbl:
        console.print(f"[yellow]配置块 '[bold]{name}[/bold]' 不存在，无需删除。[/yellow]")
        sys.exit(1)

    del llm_tbl[name]

    with open(cfg_path, "w", encoding="utf-8") as f:
        tomlkit.dump(doc, f)

    # 重新加载配置
    config._initialized = False
    config.__init__()

    console.print(f"[green]✓[/green] 已删除 LLM 配置：[bold cyan][llm.{name}][/bold cyan]")


# ── init 命令 ─────────────────────────────────────────────────────────────────

@main.command()
@click.option(
    "--force", "-f",
    is_flag=True,
    help="覆盖已有配置文件",
)
def init(force: bool):
    """初始化配置文件（首次安装或 pipx 安装后运行）。

    \b
    配置文件位置：
      macOS/Linux: ~/.config/product-search/config.toml
      Windows:     %APPDATA%\\product-search\\config.toml
    """
    from product_search.core.config import DEFAULT_CONFIG_TEMPLATE, get_user_config_dir

    cfg_dir = get_user_config_dir()
    cfg_file = cfg_dir / "config.toml"

    if cfg_file.exists() and not force:
        console.print(f"[yellow]配置文件已存在：{cfg_file}[/yellow]")
        console.print("使用 [bold]--force[/bold] 覆盖，或直接编辑该文件。")
        return

    cfg_dir.mkdir(parents=True, exist_ok=True)
    cfg_file.write_text(DEFAULT_CONFIG_TEMPLATE, encoding="utf-8")

    console.print(f"[green]✓[/green] 配置文件已创建：[bold cyan]{cfg_file}[/bold cyan]")
    console.print("\n[bold]下一步：[/bold]编辑配置文件，填入你的 API Key：")
    console.print(f"  [dim]{cfg_file}[/dim]")
    console.print("\n或通过环境变量设置（无需修改配置文件）：")
    console.print("  [dim]export OPENAI_API_KEY=sk-...[/dim]")
    console.print("  [dim]export ANTHROPIC_API_KEY=sk-ant-...[/dim]")


# ── completion 命令 ───────────────────────────────────────────────────────────

@main.command(name="completion")
@click.argument("shell", type=click.Choice(["bash", "zsh", "fish", "powershell"]))
def shell_completion(shell: str):
    """输出 shell 自动补全脚本。

    \b
    安装方法：
      bash:        product-search completion bash >> ~/.bashrc
      zsh:         product-search completion zsh >> ~/.zshrc
      fish:        product-search completion fish > ~/.config/fish/completions/product-search.fish
      powershell:  product-search completion powershell >> $PROFILE
    """
    import os
    env_var = f"_{('product_search').upper()}_COMPLETE"
    shell_env = {
        "bash": "bash_source",
        "zsh": "zsh_source",
        "fish": "fish_source",
        "powershell": "powershell_source",
    }
    os.environ[env_var] = shell_env[shell]
    try:
        main(standalone_mode=False)
    except SystemExit:
        pass


if __name__ == "__main__":
    main()
