"""交互式 REPL 模式，类 Claude CLI 体验。"""

from __future__ import annotations

import asyncio
import os
import sys
from pathlib import Path
from typing import List, Optional, Tuple

from prompt_toolkit import PromptSession
from prompt_toolkit.auto_suggest import AutoSuggestFromHistory
from prompt_toolkit.completion import WordCompleter
from prompt_toolkit.formatted_text import HTML
from prompt_toolkit.history import InMemoryHistory
from prompt_toolkit.key_binding import KeyBindings
from prompt_toolkit.patch_stdout import patch_stdout
from prompt_toolkit.styles import Style
from rich.console import Console
from rich.panel import Panel
from rich.table import Table

# ── 常量 ────────────────────────────────────────────────────────────────────

SLASH_COMMANDS = [
    "/help", "/history", "/clear",
    "/export", "/batch",
    "/quit", "/exit",
]

PROMPT_STYLE = Style.from_dict({
    "prompt":   "#00afff bold",
    "cmd":      "#00ff87",
    "dim":      "#666666",
})

WELCOME = """\
[bold cyan]ProductSearch[/bold cyan] [dim]interactive mode[/dim]
[dim]输入企业名称搜索产品，/help 查看所有命令，Ctrl+D 退出[/dim]
"""

HELP_TEXT = """\
[bold]可用命令[/bold]

  [cyan]<企业名称>[/cyan]                 搜索该企业的产品信息
  [cyan]/batch <文件路径>[/cyan]          从 Excel 批量搜索并导出
  [cyan]/export [格式] [文件名][/cyan]    导出上次结果（格式: excel/json/text，默认 excel）
  [cyan]/history[/cyan]                  查看本次会话的搜索记录
  [cyan]/clear[/cyan]                    清空屏幕
  [cyan]/help[/cyan]                     显示此帮助
  [cyan]/quit[/cyan]  或  [cyan]Ctrl+D[/cyan]         退出

[bold]快捷键[/bold]

  [dim]↑ / ↓[/dim]          翻阅历史输入
  [dim]Ctrl+C[/dim]          取消当前搜索
  [dim]Ctrl+D[/dim]          退出
"""


# ── REPL 会话 ────────────────────────────────────────────────────────────────

class ReplSession:
    """保存单次 REPL 运行期间的所有状态。"""

    def __init__(
        self,
        console: Console,
        max_iterations: int = 3,
        llm_config: str = "default",
        no_summary: bool = False,
    ):
        self.console = console
        self.max_iterations = max_iterations
        self.llm_config = llm_config
        self.no_summary = no_summary

        # 历史：[(company, product_count, state)]
        self._history: List[Tuple[str, int]] = []
        self._last_state = None
        self._last_company: Optional[str] = None

        # LLM（延迟初始化，首次搜索时建立）
        self._llm = None
        self._analysis_llm = None

    # ── LLM 初始化 ───────────────────────────────────────────────────────────

    async def _ensure_llm(self) -> bool:
        if self._llm is not None:
            return True
        try:
            from product_search.llm.factory import create_llm
            self._llm = create_llm(self.llm_config)
            if not self.no_summary:
                try:
                    self._analysis_llm = create_llm("analysis")
                except Exception:
                    self._analysis_llm = self._llm
            return True
        except Exception as e:
            self.console.print(f"[red]LLM 初始化失败: {e}[/red]")
            self.console.print(
                "[yellow]提示：请检查配置文件，或设置 OPENAI_API_KEY / ANTHROPIC_API_KEY[/yellow]"
            )
            return False

    # ── 搜索 ─────────────────────────────────────────────────────────────────

    async def search(self, company: str) -> None:
        if not await self._ensure_llm():
            return

        from product_search.graph.workflow import run_search
        from rich.progress import Progress, SpinnerColumn, TextColumn, TimeElapsedColumn

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            TimeElapsedColumn(),
            console=self.console,
            transient=True,
        ) as progress:
            progress.add_task(f"正在搜索 [cyan]{company}[/cyan]...", total=None)
            try:
                state = await run_search(
                    company_name=company,
                    max_iterations=self.max_iterations,
                    llm=self._llm,
                    analysis_llm=self._analysis_llm if not self.no_summary else None,
                )
            except asyncio.CancelledError:
                self.console.print("\n[yellow]已取消[/yellow]")
                return
            except Exception as e:
                self.console.print(f"[red]搜索失败: {e}[/red]")
                return

        products = state.get("products", [])
        summary = state.get("summary", "")

        if not products:
            self.console.print(f"[yellow]未找到 {company!r} 的产品信息[/yellow]")
            return

        self._last_state = state
        self._last_company = company
        self._history.append((company, len(products)))

        from product_search.cli import _output_table
        _output_table(products, summary, company)

    # ── 批量搜索 ─────────────────────────────────────────────────────────────

    async def batch(self, file_path: str) -> None:
        path = Path(file_path.strip().strip('"').strip("'")).expanduser().resolve()
        if not path.exists():
            self.console.print(f"[red]文件不存在: {path}[/red]")
            return

        from product_search.cli import _run_batch
        await _run_batch(
            str(path), None, None, None,
            True, self.max_iterations, self.llm_config, self.no_summary,
        )

    # ── 导出 ─────────────────────────────────────────────────────────────────

    def export(self, fmt: str = "excel", filename: Optional[str] = None) -> None:
        if self._last_state is None:
            self.console.print("[yellow]尚无搜索结果可导出，请先搜索一家企业[/yellow]")
            return

        products = self._last_state.get("products", [])
        summary = self._last_state.get("summary", "")
        company = self._last_company or "output"

        if fmt == "json":
            import json
            fname = filename or f"{company}_产品.json"
            data = {
                "company_name": company,
                "product_count": len(products),
                "products": [p.model_dump() for p in products],
                "summary": summary,
            }
            Path(fname).write_text(
                json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            self.console.print(f"[green]✓[/green] 已导出 JSON：[bold]{fname}[/bold]")

        elif fmt == "text":
            fname = filename or f"{company}_产品.txt"
            lines = [f"{company} 产品列表（共 {len(products)} 个）", "=" * 50]
            for p in products:
                lines.append(f"• {p.name}（{p.category}）")
                if p.description:
                    lines.append(f"  {p.description}")
            if summary:
                lines += ["\nAI 分析报告：", "-" * 50, summary]
            Path(fname).write_text("\n".join(lines), encoding="utf-8")
            self.console.print(f"[green]✓[/green] 已导出文本：[bold]{fname}[/bold]")

        else:  # excel
            try:
                import openpyxl
                from openpyxl.styles import Alignment, Font, PatternFill
                from openpyxl.utils import get_column_letter
            except ImportError:
                self.console.print("[red]缺少 openpyxl：pip install openpyxl[/red]")
                return

            fname = filename or f"{company}_产品.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "产品列表"

            hf = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
            hfill = PatternFill(fill_type="solid", fgColor="2E75B6")
            ha = Alignment(horizontal="center", vertical="center")
            headers = ["产品名称", "产品类别", "产品描述", "置信度", "来源URL"]
            widths  = [24, 16, 45, 8, 40]

            for col, (h, w) in enumerate(zip(headers, widths), 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font, cell.fill, cell.alignment = hf, hfill, ha
                ws.column_dimensions[get_column_letter(col)].width = w

            ws.freeze_panes = "A2"
            ca = Alignment(vertical="top", wrap_text=True)
            for row, p in enumerate(products, 2):
                ws.cell(row=row, column=1, value=p.name).alignment = ca
                ws.cell(row=row, column=2, value=p.category).alignment = ca
                ws.cell(row=row, column=3, value=p.description).alignment = ca
                ws.cell(row=row, column=4, value=round(p.confidence, 2)).alignment = ca
                ws.cell(row=row, column=5, value=p.source_url).alignment = ca

            if summary:
                ws2 = wb.create_sheet("AI汇总")
                ws2.column_dimensions["A"].width = 100
                ws2.cell(row=1, column=1, value=summary).alignment = Alignment(wrap_text=True)

            wb.save(fname)
            self.console.print(f"[green]✓[/green] 已导出 Excel：[bold]{fname}[/bold]")

    # ── 历史 ─────────────────────────────────────────────────────────────────

    def show_history(self) -> None:
        if not self._history:
            self.console.print("[dim]本次会话暂无搜索记录[/dim]")
            return
        table = Table(title="搜索历史", show_lines=False, box=None, show_header=True)
        table.add_column("#",     style="dim", width=4)
        table.add_column("企业",  style="bold cyan")
        table.add_column("产品数", justify="right", style="green")
        for i, (company, count) in enumerate(self._history, 1):
            table.add_row(str(i), company, str(count))
        self.console.print(table)


# ── 命令分发 ─────────────────────────────────────────────────────────────────

async def _dispatch(line: str, session: ReplSession) -> bool:
    """解析并执行一行输入，返回 False 表示退出。"""
    line = line.strip()
    if not line:
        return True

    lower = line.lower()

    # 退出
    if lower in ("/quit", "/exit"):
        return False

    # 帮助
    if lower == "/help":
        session.console.print(Panel(HELP_TEXT, title="帮助", border_style="blue", padding=(0, 2)))
        return True

    # 历史
    if lower == "/history":
        session.show_history()
        return True

    # 清屏（跨平台）
    if lower == "/clear":
        os.system("cls" if sys.platform == "win32" else "clear")
        return True

    # 批量
    if lower.startswith("/batch"):
        parts = line.split(maxsplit=1)
        if len(parts) < 2:
            session.console.print("[yellow]用法: /batch <Excel文件路径>[/yellow]")
        else:
            await session.batch(parts[1])
        return True

    # 导出
    if lower.startswith("/export"):
        parts = line.split()
        fmt      = parts[1] if len(parts) > 1 else "excel"
        filename = parts[2] if len(parts) > 2 else None
        if fmt not in ("excel", "json", "text"):
            session.console.print("[yellow]格式须为 excel / json / text[/yellow]")
        else:
            session.export(fmt, filename)
        return True

    # 未知斜杠命令
    if line.startswith("/"):
        session.console.print(f"[yellow]未知命令 {line!r}，输入 /help 查看帮助[/yellow]")
        return True

    # 普通文本 → 搜索
    task = asyncio.create_task(session.search(line))
    try:
        await task
    except asyncio.CancelledError:
        task.cancel()
        session.console.print("\n[yellow]已取消[/yellow]")

    return True


# ── 主入口 ───────────────────────────────────────────────────────────────────

async def run_repl(
    max_iterations: int = 3,
    llm_config: str = "default",
    no_summary: bool = False,
) -> None:
    """启动交互式 REPL。"""
    console = Console()

    # Windows 需要 ProactorEventLoop 以支持子进程（httpx 等）
    if sys.platform == "win32":
        asyncio.get_event_loop_policy().get_event_loop()

    repl = ReplSession(console, max_iterations, llm_config, no_summary)

    completer = WordCompleter(SLASH_COMMANDS, pattern=r"\/\w*", sentence=True)
    prompt_session: PromptSession = PromptSession(
        history=InMemoryHistory(),
        auto_suggest=AutoSuggestFromHistory(),
        completer=completer,
        complete_while_typing=False,
        style=PROMPT_STYLE,
        mouse_support=False,
    )

    console.print(Panel(WELCOME, border_style="cyan", padding=(0, 2)))

    with patch_stdout(raw=True):
        while True:
            try:
                line: str = await prompt_session.prompt_async(
                    HTML("<prompt>› </prompt>"),
                )
            except KeyboardInterrupt:
                # Ctrl+C 在提示符处 → 清空当前行，继续
                continue
            except EOFError:
                # Ctrl+D → 退出
                console.print("\n[dim]再见！[/dim]")
                break

            try:
                should_continue = await _dispatch(line, repl)
            except KeyboardInterrupt:
                console.print("\n[yellow]已取消[/yellow]")
                continue

            if not should_continue:
                console.print("[dim]再见！[/dim]")
                break
